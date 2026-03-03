import logging
import os
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ConversationHandler,
    ContextTypes,
    PicklePersistence,
    filters
)
from datetime import datetime, timedelta
import asyncio
import matplotlib
matplotlib.use('Agg')  # Для работы на сервере без GUI
import matplotlib.pyplot as plt
import io

import config
from database import Database

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Состояния для ConversationHandler
WAITING_FOR_INTEREST, WAITING_FOR_RELEVANCE, WAITING_FOR_SPIRITUAL, WAITING_FOR_FEEDBACK = range(4)

# Инициализация базы данных
db = Database()

# user_ratings теперь хранится в context.user_data['rating'] для persistence


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start"""
    user = update.effective_user
    user_id = user.id
    
    # Проверяем статус пользователя
    if db.is_user_approved(user_id):
        await update.message.reply_text(
            f"Привіт, {user.first_name}! Ти вже затверджений і можеш користуватися ботом.\n\n"
            "Після кожної молодіжки тобі прийде опитування для оцінки зустрічі."
        )
    elif db.is_user_pending(user_id):
        await update.message.reply_text(
            "Твій запит вже відправлено адміністратору. Очікуй затвердження!"
        )
    else:
        # Добавляем в очередь на одобрение
        db.add_pending_user(
            user_id=user_id,
            username=user.username or "",
            first_name=user.first_name or "",
            last_name=user.last_name or ""
        )
        
        await update.message.reply_text(
            "Запит на доступ відправлено адміністратору. Очікуй затвердження!"
        )
        
        # Уведомляем админа
        try:
            await context.bot.send_message(
                chat_id=config.ADMIN_ID,
                text=f"🔔 Новий запит на доступ:\n\n"
                     f"Ім'я: {user.first_name} {user.last_name or ''}\n"
                     f"Username: @{user.username or 'не вказано'}\n"
                     f"ID: {user_id}\n\n"
                     f"Використай /pending щоб переглянути всі запити."
            )
        except Exception as e:
            logger.error(f"Error notifying admin: {e}")


async def admin_pending(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает список пользователей ожидающих одобрения (только для админа)"""
    if update.effective_user.id != config.ADMIN_ID:
        await update.message.reply_text("У тебе немає доступу до цієї команди.")
        return
    
    pending_users = db.get_pending_users()
    
    if not pending_users:
        await update.message.reply_text("Немає користувачів, що очікують затвердження.")
        return
    
    for user in pending_users:
        user_id, username, first_name, last_name, request_date = user
        keyboard = [
            [
                InlineKeyboardButton("✅ Затвердити", callback_data=f"approve_{user_id}"),
                InlineKeyboardButton("❌ Відхилити", callback_data=f"reject_{user_id}")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            f"👤 Запит:\n"
            f"Ім'я: {first_name} {last_name or ''}\n"
            f"Username: @{username or 'не вказано'}\n"
            f"ID: {user_id}\n"
            f"Дата запиту: {request_date[:16]}",
            reply_markup=reply_markup
        )


async def admin_remove(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает список одобренных пользователей для удаления (только для админа)"""
    if update.effective_user.id != config.ADMIN_ID:
        await update.message.reply_text("У тебе немає доступу до цієї команди.")
        return
    
    approved_users = db.get_all_approved_users_info()
    
    if not approved_users:
        await update.message.reply_text("Немає затверджених користувачів.")
        return
    
    # Фильтруем админа из списка
    approved_users = [u for u in approved_users if u[0] != config.ADMIN_ID]
    
    if not approved_users:
        await update.message.reply_text("Немає користувачів для видалення (крім тебе).")
        return
    
    text = "👥 *Затверджені користувачі:*\n\n"
    text += "Виберь користувача для видалення:\n\n"
    
    for user_id, username, first_name, last_name in approved_users:
        keyboard = [
            [InlineKeyboardButton("🗑 Видалити", callback_data=f"remove_{user_id}")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            f"👤 Користувач:\n"
            f"Ім'я: {first_name} {last_name or ''}\n"
            f"Username: @{username or 'не вказано'}\n"
            f"ID: {user_id}",
            reply_markup=reply_markup
        )


async def handle_approval(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик кнопок одобрения/отклонения/удаления"""
    query = update.callback_query
    await query.answer()
    
    if query.from_user.id != config.ADMIN_ID:
        await query.edit_message_text("У тебе немає доступу до цієї дії.")
        return
    
    action, user_id = query.data.split('_')
    user_id = int(user_id)
    
    if action == "approve":
        db.approve_user(user_id)
        await query.edit_message_text(f"✅ Користувача {user_id} затверджено!")
        
        # Уведомляем пользователя
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text="🎉 Твій запит затверджено! Тепер ти будеш отримувати опитування після молодіжних зустрічей."
            )
            
            # Проверяем есть ли активное опитування
            active_meeting = db.get_active_meeting()
            if active_meeting:
                # Отправляем активное опитування новому пользователю
                keyboard = [
                    [InlineKeyboardButton("📝 Оцінити", callback_data=f"rate_{active_meeting}")],
                    [InlineKeyboardButton("❌ Не був на молодіжці", callback_data=f"absent_{active_meeting}")]
                ]
                reply_markup = InlineKeyboardMarkup(keyboard)
                
                await context.bot.send_message(
                    chat_id=user_id,
                    text="🙏 Привіт! Будь ласка, оціни минулу молодіжку.\n\n"
                         f"У тебе є {config.RATING_DEADLINE_HOURS} годин на оцінку.\n"
                         "За годину до закінчення прийде нагадування.",
                    reply_markup=reply_markup
                )
                
                # Регистрируем пользователя для этой встречи
                db.register_user_for_meeting(active_meeting, user_id)
                
                logger.info(f"Sent active survey {active_meeting} to newly approved user {user_id}")
        except Exception as e:
            logger.error(f"Error notifying approved user: {e}")
    
    elif action == "reject":
        db.reject_user(user_id)
        await query.edit_message_text(f"❌ Запит користувача {user_id} відхилено.")
        
        # Уведомляем пользователя
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text="На жаль, твій запит на доступ було відхилено."
            )
        except Exception as e:
            logger.error(f"Error notifying rejected user: {e}")
    
    elif action == "remove":
        if db.remove_user(user_id):
            await query.edit_message_text(f"🗑 Користувача {user_id} видалено зі списку!")
            
            # Уведомляем пользователя
            try:
                await context.bot.send_message(
                    chat_id=user_id,
                    text="Тебе було видалено зі списку учасників бота. Ти більше не будеш отримувати опитування."
                )
            except Exception as e:
                logger.error(f"Error notifying removed user: {e}")
        else:
            await query.edit_message_text(f"❌ Користувача {user_id} не знайдено.")


async def admin_start_survey(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Запускает новый опрос (только для админа)"""
    if update.effective_user.id != config.ADMIN_ID:
        await update.message.reply_text("У тебе немає доступу до цієї команди.")
        return
    
    # Проверяем нет ли активного опроса
    active_meeting = db.get_active_meeting()
    if active_meeting:
        await update.message.reply_text(
            "❌ Вже є активне опитування! Спочатку дочекайся його завершення або закрий його командою /close_survey"
        )
        return
    
    # Создаем новую встречу
    meeting_id = db.create_meeting()
    
    # Рассылаем опрос всем одобренным пользователям
    approved_users = db.get_all_approved_users()
    
    if not approved_users:
        await update.message.reply_text("❌ Немає затверджених користувачів для опитування!")
        return
    
    success_count = 0
    for user_id in approved_users:
        if user_id == config.ADMIN_ID:
            continue  # Не отправляем админу
        
        try:
            keyboard = [
                [InlineKeyboardButton("📝 Оцінити", callback_data=f"rate_{meeting_id}")],
                [InlineKeyboardButton("❌ Не був на молодіжці", callback_data=f"absent_{meeting_id}")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await context.bot.send_message(
                chat_id=user_id,
                text="🙏 Привіт! Будь ласка, оціни минулу молодіжку.\n\n"
                     f"У тебе є {config.RATING_DEADLINE_HOURS} годин на оцінку.\n"
                     "За годину до закінчення прийде нагадування.",
                reply_markup=reply_markup
            )
            success_count += 1
        except Exception as e:
            logger.error(f"Error sending survey to user {user_id}: {e}")
    
    await update.message.reply_text(
        f"✅ Опитування запущено! ID зустрічі: {meeting_id}\n"
        f"Відправлено {success_count} користувачам.\n\n"
        f"Дедлайн: {config.RATING_DEADLINE_HOURS} годин\n"
        f"Нагадування буде відправлено за {config.REMINDER_BEFORE_DEADLINE_HOURS} годину до кінця."
    )
    
    # Планируем напоминание и закрытие опроса
    reminder_time = config.RATING_DEADLINE_HOURS - config.REMINDER_BEFORE_DEADLINE_HOURS
    
    # Запланируем джобы
    context.job_queue.run_once(
        send_reminders,
        reminder_time * 3600,
        data={'meeting_id': meeting_id},
        name=f'reminder_{meeting_id}'
    )
    
    context.job_queue.run_once(
        close_survey_job,
        config.RATING_DEADLINE_HOURS * 3600,
        data={'meeting_id': meeting_id},
        name=f'close_{meeting_id}'
    )


async def send_reminders(context: ContextTypes.DEFAULT_TYPE):
    """Отправляет напоминания тем, кто еще не оценил"""
    meeting_id = context.job.data['meeting_id']
    users_to_remind = db.get_users_for_reminder(meeting_id)
    
    for user_id in users_to_remind:
        if user_id == config.ADMIN_ID:
            continue
        
        try:
            keyboard = [
                [InlineKeyboardButton("📝 Оцінити", callback_data=f"rate_{meeting_id}")],
                [InlineKeyboardButton("❌ Не був на молодіжці", callback_data=f"absent_{meeting_id}")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await context.bot.send_message(
                chat_id=user_id,
                text=f"⏰ Нагадування: у тебе залишилася {config.REMINDER_BEFORE_DEADLINE_HOURS} година щоб оцінити молодіжку!\n\n"
                     "Будь ласка, не забудь залишити зворотний зв'язок.",
                reply_markup=reply_markup
            )
            db.mark_as_reminded(meeting_id, user_id)
        except Exception as e:
            logger.error(f"Error sending reminder to user {user_id}: {e}")


async def close_survey_job(context: ContextTypes.DEFAULT_TYPE):
    """Автоматически закрывает опрос по истечении времени"""
    meeting_id = context.job.data['meeting_id']
    db.close_meeting(meeting_id)
    
    # Уведомляем админа
    try:
        stats = db.get_meeting_stats(meeting_id)
        await context.bot.send_message(
            chat_id=config.ADMIN_ID,
            text=f"⏱ Опитування #{meeting_id} автоматично закрито.\n\n"
                 f"Використай /stats {meeting_id} щоб переглянути результати."
        )
    except Exception as e:
        logger.error(f"Error notifying admin about closed survey: {e}")


async def admin_close_survey(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Вручную закрывает активный опрос (только для админа)"""
    if update.effective_user.id != config.ADMIN_ID:
        await update.message.reply_text("У тебе немає доступу до цієї команди.")
        return
    
    active_meeting = db.get_active_meeting()
    if not active_meeting:
        await update.message.reply_text("❌ Немає активного опитування.")
        return
    
    db.close_meeting(active_meeting)
    
    # Отменяем запланированные джобы
    current_jobs = context.job_queue.get_jobs_by_name(f'reminder_{active_meeting}')
    for job in current_jobs:
        job.schedule_removal()
    
    current_jobs = context.job_queue.get_jobs_by_name(f'close_{active_meeting}')
    for job in current_jobs:
        job.schedule_removal()
    
    await update.message.reply_text(
        f"✅ Опитування #{active_meeting} закрито вручну.\n\n"
        f"Використай /stats {active_meeting} щоб переглянути результати."
    )


async def handle_rating_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик кнопки 'Оценить'"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    
    # Проверяем что пользователь одобрен
    if not db.is_user_approved(user_id):
        await query.edit_message_text("У тебе немає доступу до цього бота.")
        return
    
    data = query.data.split('_')
    action = data[0]
    meeting_id = int(data[1])
    
    if action == "absent":
        # Пользователь не был на встрече
        db.mark_not_attended(meeting_id, user_id)
        await query.edit_message_text(
            "✅ Дякуємо за відповідь! Сподіваємося побачити тебе на наступній молодіжці! 🙏"
        )
        return
    
    elif action == "rate":
        # Начинаем процесс оценки - сохраняем в context.user_data для persistence
        context.user_data['rating'] = {
            'meeting_id': meeting_id,
            'interest': None,
            'relevance': None,
            'spiritual': None
        }
        
        keyboard = [
            [InlineKeyboardButton(str(i), callback_data=f"interest_{i}") for i in range(1, 6)]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            "📊 Оціни *цікавість* молодіжки від 1 до 5:\n\n"
            "1 - Нудно\n"
            "5 - Дуже цікаво",
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
        return WAITING_FOR_INTEREST


async def handle_interest_rating(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик оценки интересности"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    rating = int(query.data.split('_')[1])

    context.user_data['rating']['interest'] = rating

    keyboard = [
        [InlineKeyboardButton(str(i), callback_data=f"relevance_{i}") for i in range(1, 6)]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        "📊 Оціни *актуальність для тебе* від 1 до 5:\n\n"
        "1 - Зовсім не актуально\n"
        "5 - Дуже актуально",
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )
    return WAITING_FOR_RELEVANCE


async def handle_relevance_rating(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик оценки актуальности"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    rating = int(query.data.split('_')[1])

    context.user_data['rating']['relevance'] = rating

    keyboard = [
        [InlineKeyboardButton(str(i), callback_data=f"spiritual_{i}") for i in range(1, 6)]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        "📊 Оціни *корисність для духовного зростання* від 1 до 5:\n\n"
        "1 - Зовсім не корисно\n"
        "5 - Дуже корисно",
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )
    return WAITING_FOR_SPIRITUAL


async def handle_spiritual_rating(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик оценки духовного роста"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    rating = int(query.data.split('_')[1])

    context.user_data['rating']['spiritual'] = rating

    keyboard = [
        [InlineKeyboardButton("✍️ Залишити відгук", callback_data="feedback_yes")],
        [InlineKeyboardButton("⏭ Пропустити", callback_data="feedback_no")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        "✅ Дякуємо за оцінки!\n\n"
        "Хочеш залишити письмовий відгук? (3-4 речення)",
        reply_markup=reply_markup
    )
    return WAITING_FOR_FEEDBACK


async def handle_feedback_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик выбора оставить отзыв или нет"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    choice = query.data.split('_')[1]
    
    if choice == "no":
        # Сохраняем оценки без отзыва
        rating_data = context.user_data.get('rating')
        if rating_data:
            db.add_rating(
                meeting_id=rating_data['meeting_id'],
                user_id=user_id,
                interest=rating_data['interest'],
                relevance=rating_data['relevance'],
                spiritual_growth=rating_data['spiritual'],
                attended=True
            )
            context.user_data.pop('rating', None)

        await query.edit_message_text(
            "✅ Дякуємо за зворотний зв'язок! 🙏"
        )
        return ConversationHandler.END
    
    else:
        # Просим написать отзыв
        await query.edit_message_text(
            "✍️ Напиши свій відгук (3-4 речення):"
        )
        # Сохраняем контекст для обработки следующего текстового сообщения
        return WAITING_FOR_FEEDBACK


async def handle_feedback_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик текстового отзыва"""
    user_id = update.effective_user.id
    feedback_text = update.message.text

    rating_data = context.user_data.get('rating')
    if not rating_data:
        await update.message.reply_text("Сталася помилка. Спробуй почати оцінювання заново.")
        return ConversationHandler.END

    # Сохраняем оценки
    db.add_rating(
        meeting_id=rating_data['meeting_id'],
        user_id=user_id,
        interest=rating_data['interest'],
        relevance=rating_data['relevance'],
        spiritual_growth=rating_data['spiritual'],
        attended=True
    )

    # Сохраняем отзыв
    db.add_feedback(rating_data['meeting_id'], feedback_text)

    context.user_data.pop('rating', None)

    await update.message.reply_text(
        "✅ Дякуємо за детальний зворотний зв'язок! 🙏"
    )
    return ConversationHandler.END


async def admin_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает статистику по встрече (только для админа)"""
    if update.effective_user.id != config.ADMIN_ID:
        await update.message.reply_text("У тебе немає доступу до цієї команди.")
        return
    
    # Получаем ID встречи из аргументов или берем последнюю активную
    if context.args:
        try:
            meeting_id = int(context.args[0])
        except ValueError:
            await update.message.reply_text("❌ Невірний формат ID зустрічі.")
            return
    else:
        # Если аргумента нет - показываем список всех встреч
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT meeting_id, start_date, is_active 
            FROM youth_meetings 
            ORDER BY start_date DESC 
            LIMIT 10
        ''')
        meetings = cursor.fetchall()
        conn.close()
        
        if not meetings:
            await update.message.reply_text("❌ Ще не було жодної молодіжки.")
            return
        
        # Формируем список встреч
        from datetime import datetime
        text = "📊 *Список молодіжних зустрічей:*\n\n"
        for meeting_id, start_date, is_active in meetings:
            date_obj = datetime.fromisoformat(start_date)
            date_str = date_obj.strftime("%d.%m.%Y %H:%M")
            status = "🟢 Активна" if is_active else "⚪️ Завершена"
            text += f"#{meeting_id} - {date_str} {status}\n"
        
        text += f"\n💡 Використай `/stats ID` щоб переглянути статистику\n"
        text += f"Наприклад: `/stats 1`"
        
        await update.message.reply_text(text, parse_mode='Markdown')
        return
    
    stats = db.get_meeting_stats(meeting_id)

    # Формируем текст статистики (основная часть)
    text = f"📊 *Статистика зустрічі #{meeting_id}*\n\n"
    text += f"👥 Були присутні: {stats['total_attended']}\n"
    text += f"❌ Не було: {stats['not_attended']}\n\n"

    if stats['total_attended'] > 0:
        text += f"⭐️ *Середні оцінки:*\n"
        text += f"Цікавість: {stats['avg_interest']}/5\n"
        text += f"Актуальність: {stats['avg_relevance']}/5\n"
        text += f"Духовне зростання: {stats['avg_spiritual_growth']}/5\n"

    # Отправляем основную статистику
    await update.message.reply_text(text, parse_mode='Markdown')

    # Отправляем отзывы отдельно (разбиваем на части если много)
    if stats['feedbacks']:
        feedbacks_text = f"💬 *Відгуки ({len(stats['feedbacks'])}):*\n\n"
        feedback_count = 0

        for i, (feedback, date) in enumerate(stats['feedbacks'], 1):
            # Обрезаем длинные отзывы
            if len(feedback) > 500:
                feedback = feedback[:500] + "..."

            new_entry = f"{i}. {feedback}\n\n"

            # Если сообщение станет слишком длинным - отправляем и начинаем новое
            if len(feedbacks_text) + len(new_entry) > 3800:
                await update.message.reply_text(feedbacks_text, parse_mode='Markdown')
                feedbacks_text = f"💬 *Відгуки (продовження):*\n\n"

            feedbacks_text += new_entry
            feedback_count += 1

        # Отправляем оставшиеся отзывы
        if feedbacks_text.strip() and feedback_count > 0:
            await update.message.reply_text(feedbacks_text, parse_mode='Markdown')
    else:
        await update.message.reply_text("💬 Текстових відгуків немає.")


async def admin_ratings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает список всех оценок по встрече в анонимном формате (только для админа)"""
    if update.effective_user.id != config.ADMIN_ID:
        await update.message.reply_text("У тебе немає доступу до цієї команди.")
        return
    
    # Получаем ID встречи из аргументов
    if not context.args:
        await update.message.reply_text(
            "Вкажи ID зустрічі.\n\n"
            "Наприклад: `/ratings 2`\n\n"
            "Щоб дізнатися ID зустрічей, використай `/stats`",
            parse_mode='Markdown'
        )
        return
    
    try:
        meeting_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Невірний формат ID зустрічі.")
        return
    
    # Получаем все оценки по встрече
    conn = db.get_connection()
    cursor = conn.cursor()
    
    # Проверяем существует ли встреча
    cursor.execute('SELECT meeting_id, start_date FROM youth_meetings WHERE meeting_id = ?', (meeting_id,))
    meeting = cursor.fetchone()
    
    if not meeting:
        await update.message.reply_text(f"❌ Зустріч #{meeting_id} не знайдено.")
        conn.close()
        return
    
    # Получаем все оценки в порядке их добавления
    cursor.execute('''
        SELECT 
            interest_rating,
            relevance_rating,
            spiritual_growth_rating,
            attended,
            rating_date
        FROM ratings
        WHERE meeting_id = ?
        ORDER BY rating_date
    ''', (meeting_id,))
    
    ratings = cursor.fetchall()
    conn.close()
    
    if not ratings:
        await update.message.reply_text(f"❌ Немає оцінок для зустрічі #{meeting_id}.")
        return
    
    # Формируем текст с анонимными оценками
    from datetime import datetime
    meeting_date = datetime.fromisoformat(meeting[1]).strftime("%d.%m.%Y %H:%M")
    
    text = f"📋 *Анонімні оцінки зустрічі #{meeting_id}*\n"
    text += f"📅 {meeting_date}\n\n"
    
    user_num = 1
    for interest, relevance, spiritual, attended, rating_date in ratings:
        if attended:
            text += f"👤 *Учасник {user_num}:*\n"
            text += f"• Цікавість: {interest}/5\n"
            text += f"• Актуальність: {relevance}/5\n"
            text += f"• Духовне зростання: {spiritual}/5\n"
            text += f"_Оцінено: {datetime.fromisoformat(rating_date).strftime('%d.%m %H:%M')}_\n\n"
            user_num += 1
        else:
            text += f"❌ *Не був присутній*\n"
            text += f"_Відмітка: {datetime.fromisoformat(rating_date).strftime('%d.%m %H:%M')}_\n\n"
    
    text += f"📊 Всього оцінок: {user_num - 1}"
    
    await update.message.reply_text(text, parse_mode='Markdown')


async def admin_graph(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Создает график динамики оценок (только для админа)"""
    if update.effective_user.id != config.ADMIN_ID:
        await update.message.reply_text("У тебе немає доступу до цієї команди.")
        return
    
    # Получаем тип графика из аргументов
    if not context.args or context.args[0] not in ['month', 'year', 'all']:
        await update.message.reply_text(
            "Вкажи тип графіка:\n"
            "📊 /graph month - за місяць (по тижнях)\n"
            "📊 /graph year - за рік (по місяцях)\n"
            "📊 /graph all - за весь період (по кварталах)"
        )
        return
    
    graph_type = context.args[0]
    
    # Получаем данные в зависимости от типа
    if graph_type == 'month':
        stats = db.get_stats_for_period(30)
        title = "Динаміка оцінок за місяць"
        group_by = 'week'
    elif graph_type == 'year':
        stats = db.get_stats_for_period(365)
        title = "Динаміка оцінок за рік"
        group_by = 'month'
    else:  # all
        stats = db.get_all_stats()
        title = "Динаміка оцінок за весь період"
        group_by = 'quarter'
    
    if not stats:
        await update.message.reply_text("❌ Немає даних за вказаний період.")
        return
    
    # Если данных меньше 2 точек, предупреждаем
    if len(stats) < 2:
        await update.message.reply_text(
            f"⚠️ Недостатньо даних для графіка (тільки {len(stats)} зустріч).\n"
            "Графік буде більш інформативним після 3+ зустрічей."
        )
    
    # Группируем данные
    from datetime import datetime
    from collections import defaultdict
    
    grouped_data = defaultdict(lambda: {'interest': [], 'relevance': [], 'spiritual': [], 'dates': []})
    
    for s in stats:
        date_obj = datetime.fromisoformat(s['date'])
        
        if group_by == 'week':
            # Группируем по неделям (понедельник каждой недели)
            week_start = date_obj - timedelta(days=date_obj.weekday())
            key = week_start.strftime('%Y-%W')
            display_date = week_start
        elif group_by == 'month':
            # Группируем по месяцам
            key = date_obj.strftime('%Y-%m')
            display_date = date_obj.replace(day=1)
        else:  # quarter
            # Группируем по кварталам
            quarter = (date_obj.month - 1) // 3 + 1
            key = f"{date_obj.year}-Q{quarter}"
            quarter_month = (quarter - 1) * 3 + 1
            display_date = date_obj.replace(month=quarter_month, day=1)
        
        grouped_data[key]['interest'].append(s['avg_interest'])
        grouped_data[key]['relevance'].append(s['avg_relevance'])
        grouped_data[key]['spiritual'].append(s['avg_spiritual'])
        grouped_data[key]['dates'].append(display_date)
    
    # Вычисляем средние по группам
    dates = []
    interest = []
    relevance = []
    spiritual = []
    overall = []  # Финальная оценка
    
    for key in sorted(grouped_data.keys()):
        data = grouped_data[key]
        dates.append(data['dates'][0])
        
        # Среднее по каждой метрике в этом периоде
        avg_interest = sum(data['interest']) / len(data['interest'])
        avg_relevance = sum(data['relevance']) / len(data['relevance'])
        avg_spiritual = sum(data['spiritual']) / len(data['spiritual'])
        
        interest.append(avg_interest)
        relevance.append(avg_relevance)
        spiritual.append(avg_spiritual)
        
        # Финальная оценка = среднее трех метрик
        overall.append((avg_interest + avg_relevance + avg_spiritual) / 3)
    
    # Создаем график
    plt.figure(figsize=(14, 7))
    plt.plot(dates, interest, marker='o', label='Цікавість', linewidth=2.5, markersize=10, color='#1f77b4')
    plt.plot(dates, relevance, marker='s', label='Актуальність', linewidth=2.5, markersize=10, color='#ff7f0e')
    plt.plot(dates, spiritual, marker='^', label='Духовне зростання', linewidth=2.5, markersize=10, color='#2ca02c')
    plt.plot(dates, overall, marker='D', label='🎯 Фінальна оцінка', linewidth=3, markersize=12, color='#d62728', linestyle='--')
    
    # Настройка осей
    if group_by == 'week':
        plt.xlabel('Тиждень', fontsize=12)
        from matplotlib.dates import WeekdayLocator, DateFormatter
        plt.gca().xaxis.set_major_locator(WeekdayLocator(byweekday=0))  # Понедельники
        plt.gca().xaxis.set_major_formatter(DateFormatter('%d.%m'))
    elif group_by == 'month':
        plt.xlabel('Місяць', fontsize=12)
        from matplotlib.dates import MonthLocator, DateFormatter
        plt.gca().xaxis.set_major_locator(MonthLocator())
        plt.gca().xaxis.set_major_formatter(DateFormatter('%b %Y'))
    else:  # quarter
        plt.xlabel('Квартал', fontsize=12)
        from matplotlib.dates import MonthLocator, DateFormatter
        plt.gca().xaxis.set_major_locator(MonthLocator(interval=3))
        plt.gca().xaxis.set_major_formatter(DateFormatter('Q%q %Y'))
    
    plt.ylabel('Оцінка (1-5)', fontsize=12)
    plt.title(title, fontsize=14, fontweight='bold')
    plt.legend(fontsize=11, loc='best')
    plt.grid(True, alpha=0.3, linestyle='--')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.ylim(0, 5.5)
    
    # Сохраняем график в BytesIO
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    plt.close()
    
    # Формируем подпись
    period_names = {'month': 'місяць', 'year': 'рік', 'all': 'весь період'}
    
    # Средние по каждой метрике за период
    final_avg = sum(overall) / len(overall) if overall else 0
    
    caption = f"📈 Графік за {period_names[graph_type]}\n"
    caption += f"📊 Кількість періодів: {len(dates)}\n\n"
    caption += f"⭐️ Середні оцінки за період:\n"
    caption += f"  • Цікавість: {sum(interest)/len(interest):.2f}/5\n"
    caption += f"  • Актуальність: {sum(relevance)/len(relevance):.2f}/5\n"
    caption += f"  • Духовне зростання: {sum(spiritual)/len(spiritual):.2f}/5\n"
    caption += f"  • 🎯 Фінальна оцінка: {final_avg:.2f}/5"
    
    # Отправляем график
    await update.message.reply_photo(
        photo=buf,
        caption=caption
    )


async def admin_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает список команд для админа"""
    if update.effective_user.id != config.ADMIN_ID:
        await update.message.reply_text(
            "Доступні команди:\n"
            "/start - Почати роботу з ботом"
        )
        return
    
    help_text = """
🤖 *Команди адміністратора:*

👥 *Управління користувачами:*
/pending - Показати запити на доступ
/remove - Видалити учасника з бота

📊 *Управління опитуваннями:*
/start\\_survey - Запустити нове опитування
/close\\_survey - Закрити активне опитування вручну

📈 *Статистика:*
/stats - Список всіх зустрічей
/stats ID - Статистика по конкретному опитуванню
/ratings ID - Анонімний список оцінок по зустрічі
/graph month - Графік за місяць (по тижнях)
/graph year - Графік за рік (по місяцях)
/graph all - Графік за весь період (по кварталах)

💾 *Експорт даних:*
/export\\_excel - Завантажити дані в Excel
/export\\_db - Завантажити базу даних SQLite

❓ /help - Показати це повідомлення
    """
    
    await update.message.reply_text(help_text, parse_mode='Markdown')


async def admin_export_db(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отправляет файл базы данных админу"""
    if update.effective_user.id != config.ADMIN_ID:
        await update.message.reply_text("У тебе немає доступу до цієї команди.")
        return
    
    db_path = config.DATABASE_NAME
    
    if not os.path.exists(db_path):
        await update.message.reply_text("❌ Файл бази даних не знайдено.")
        return
    
    # Получаем статистику по базе
    conn = db.get_connection()
    cursor = conn.cursor()
    
    # Количество пользователей
    cursor.execute('SELECT COUNT(*) FROM users')
    users_count = cursor.fetchone()[0]
    
    # Количество встреч
    cursor.execute('SELECT COUNT(*) FROM youth_meetings')
    meetings_count = cursor.fetchone()[0]
    
    # Количество оценок
    cursor.execute('SELECT COUNT(*) FROM ratings WHERE attended = 1')
    ratings_count = cursor.fetchone()[0]
    
    # Количество отзывов
    cursor.execute('SELECT COUNT(*) FROM feedback')
    feedback_count = cursor.fetchone()[0]
    
    # Размер файла
    file_size = os.path.getsize(db_path)
    file_size_mb = file_size / 1024 / 1024
    
    conn.close()
    
    # Формируем описание
    caption = f"💾 *База даних*\n\n"
    caption += f"👥 Користувачів: {users_count}\n"
    caption += f"📅 Зустрічей: {meetings_count}\n"
    caption += f"⭐️ Оцінок: {ratings_count}\n"
    caption += f"💬 Відгуків: {feedback_count}\n"
    caption += f"📦 Розмір: {file_size_mb:.2f} МБ\n\n"
    caption += f"🔧 Відкрити можна за допомогою SQLite Browser або будь-якого SQL клієнта"
    
    # Отправляем файл
    await update.message.reply_document(
        document=open(db_path, 'rb'),
        filename=f'youth_feedback_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db',
        caption=caption,
        parse_mode='Markdown'
    )


async def auto_backup(context: ContextTypes.DEFAULT_TYPE):
    """Автоматичний щотижневий бекап бази даних"""
    try:
        db_path = config.DATABASE_NAME

        if not os.path.exists(db_path):
            logger.error("Auto backup: database file not found")
            return

        # Отримуємо статистику
        conn = db.get_connection()
        cursor = conn.cursor()

        cursor.execute('SELECT COUNT(*) FROM users')
        users_count = cursor.fetchone()[0]

        cursor.execute('SELECT COUNT(*) FROM youth_meetings')
        meetings_count = cursor.fetchone()[0]

        cursor.execute('SELECT COUNT(*) FROM ratings WHERE attended = 1')
        ratings_count = cursor.fetchone()[0]

        file_size = os.path.getsize(db_path)
        file_size_kb = file_size / 1024

        conn.close()

        caption = f"🔄 *Автоматичний бекап*\n\n"
        caption += f"👥 Користувачів: {users_count}\n"
        caption += f"📅 Зустрічей: {meetings_count}\n"
        caption += f"⭐️ Оцінок: {ratings_count}\n"
        caption += f"📦 Розмір: {file_size_kb:.1f} КБ\n\n"
        caption += f"📆 {datetime.now().strftime('%d.%m.%Y %H:%M')}"

        await context.bot.send_document(
            chat_id=config.ADMIN_ID,
            document=open(db_path, 'rb'),
            filename=f'backup_{datetime.now().strftime("%Y%m%d")}.db',
            caption=caption,
            parse_mode='Markdown'
        )
        logger.info("Auto backup sent successfully")

    except Exception as e:
        logger.error(f"Auto backup error: {e}")


async def check_and_close_expired_surveys(context: ContextTypes.DEFAULT_TYPE):
    """Фонова задача: перевіряє і закриває прострочені опитування"""
    try:
        active_meeting = db.get_active_meeting()
        if not active_meeting:
            return
        
        # Получаем дедлайн встречи
        deadline = db.get_meeting_deadline(active_meeting)
        if not deadline:
            return
        
        # Проверяем истек ли дедлайн
        now = datetime.now()
        if now >= deadline:
            logger.info(f"Auto-closing expired survey {active_meeting}")
            
            # Закрываем встречу
            db.close_meeting(active_meeting)
            
            # Получаем статистику
            stats = db.get_meeting_stats(active_meeting)
            
            # Формируем сообщение для админа
            text = f"⏰ *Опитування #{active_meeting} автоматично закрито*\n\n"
            text += f"📊 *Підсумки:*\n"
            text += f"👥 Відповіли: {stats['total_attended']}\n"
            text += f"❌ Не було: {stats['not_attended']}\n\n"
            
            if stats['total_attended'] > 0:
                text += f"⭐️ *Середні оцінки:*\n"
                text += f"• Цікавість: {stats['avg_interest']}/5\n"
                text += f"• Актуальність: {stats['avg_relevance']}/5\n"
                text += f"• Духовне зростання: {stats['avg_spiritual_growth']}/5\n\n"
            
            text += f"💡 Використай `/stats {active_meeting}` для детальної статистики"
            
            # Отправляем админу
            await context.bot.send_message(
                chat_id=config.ADMIN_ID,
                text=text,
                parse_mode='Markdown'
            )
            
            logger.info(f"Survey {active_meeting} auto-closed and admin notified")
            
    except Exception as e:
        logger.error(f"Error in check_and_close_expired_surveys: {e}")


async def admin_export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспортирует базу данных в Excel"""
    if update.effective_user.id != config.ADMIN_ID:
        await update.message.reply_text("У тебе немає доступу до цієї команди.")
        return
    
    await update.message.reply_text("⏳ Створюю Excel файл...")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        # Удаляем стандартный лист
        wb.remove(wb.active)
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # === ЛИСТ 1: Зустрічі ===
        ws_meetings = wb.create_sheet("Зустрічі")
        ws_meetings.append(["ID", "Дата початку", "Активна", "Середня цікавість", "Середня актуальність", "Середнє духовне зростання", "Відвідали"])
        
        cursor.execute('''
            SELECT 
                m.meeting_id,
                m.start_date,
                CASE WHEN m.is_active = 1 THEN 'Так' ELSE 'Ні' END,
                ROUND(AVG(CASE WHEN r.attended = 1 THEN r.interest_rating END), 2),
                ROUND(AVG(CASE WHEN r.attended = 1 THEN r.relevance_rating END), 2),
                ROUND(AVG(CASE WHEN r.attended = 1 THEN r.spiritual_growth_rating END), 2),
                COUNT(CASE WHEN r.attended = 1 THEN 1 END)
            FROM youth_meetings m
            LEFT JOIN ratings r ON m.meeting_id = r.meeting_id
            GROUP BY m.meeting_id
            ORDER BY m.start_date DESC
        ''')
        for row in cursor.fetchall():
            ws_meetings.append(list(row))
        
        for cell in ws_meetings[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # === ЛИСТ 3: Оцінки ===
        ws_ratings = wb.create_sheet("Оцінки")
        ws_ratings.append(["ID зустрічі", "Дата зустрічі", "Відвідав", "Цікавість", "Актуальність", "Духовне зростання", "Дата оцінки"])
        
        cursor.execute('''
            SELECT 
                m.meeting_id,
                m.start_date,
                CASE WHEN r.attended = 1 THEN 'Так' ELSE 'Ні' END,
                r.interest_rating,
                r.relevance_rating,
                r.spiritual_growth_rating,
                r.rating_date
            FROM ratings r
            JOIN youth_meetings m ON r.meeting_id = m.meeting_id
            ORDER BY m.start_date DESC, r.rating_date
        ''')
        for row in cursor.fetchall():
            ws_ratings.append(list(row))
        
        for cell in ws_ratings[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # === ЛИСТ 4: Відгуки ===
        ws_feedback = wb.create_sheet("Відгуки")
        ws_feedback.append(["ID зустрічі", "Дата зустрічі", "Відгук", "Дата відгуку"])
        
        cursor.execute('''
            SELECT 
                m.meeting_id,
                m.start_date,
                f.feedback_text,
                f.feedback_date
            FROM feedback f
            JOIN youth_meetings m ON f.meeting_id = m.meeting_id
            ORDER BY m.start_date DESC, f.feedback_date
        ''')
        for row in cursor.fetchall():
            ws_feedback.append(list(row))
        
        for cell in ws_feedback[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="9966FF", end_color="9966FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Устанавливаем ширину колонок
        for ws in [ws_meetings, ws_ratings, ws_feedback]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        conn.close()
        
        # Сохраняем файл
        filename = f'youth_feedback_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(filename)
        
        # Получаем статистику
        file_size = os.path.getsize(filename)
        file_size_kb = file_size / 1024
        
        cursor = db.get_connection().cursor()
        cursor.execute('SELECT COUNT(*) FROM users')
        users_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM youth_meetings')
        meetings_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM ratings WHERE attended = 1')
        ratings_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM feedback')
        feedback_count = cursor.fetchone()[0]
        
        # Формируем описание
        caption = f"📊 *Excel експорт бази даних*\n\n"
        caption += f"📋 Листи:\n"
        caption += f"• Зустрічі ({meetings_count})\n"
        caption += f"• Оцінки ({ratings_count})\n"
        caption += f"• Відгуки ({feedback_count})\n\n"
        caption += f"👥 Користувачів в системі: {users_count}\n"
        caption += f"📦 Розмір: {file_size_kb:.1f} КБ\n"
        caption += f"🗓 Створено: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        # Отправляем файл
        await update.message.reply_document(
            document=open(filename, 'rb'),
            filename=filename,
            caption=caption,
            parse_mode='Markdown'
        )
        
        # Удаляем временный файл
        os.remove(filename)
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        await update.message.reply_text(f"❌ Помилка при створенні Excel файлу: {str(e)}")


def main():
    """Главная функция запуска бота"""
    # Проверяем что ADMIN_ID установлен
    if config.ADMIN_ID == 0:
        logger.error("ADMIN_ID not set! Please set your Telegram user ID in config.py")
        return

    # Persistence - сохраняет состояние ConversationHandler и user_data между перезапусками
    # Удаляем старый pickle при старте чтобы сбросить застрявшие состояния диалогов
    pickle_path = "/var/data/bot_persistence.pickle"
    if os.path.exists(pickle_path):
        os.remove(pickle_path)
        logger.info("Cleared old persistence file to reset stuck conversation states")
    persistence = PicklePersistence(filepath=pickle_path)
    logger.info("Persistence enabled - state will be saved to " + pickle_path)

    # Создаем приложение с persistence
    application = Application.builder().token(config.BOT_TOKEN).persistence(persistence).build()
    
    # Добавляем фоновую задачу проверки дедлайнов (каждую 1 час)
    job_queue = application.job_queue
    job_queue.run_repeating(check_and_close_expired_surveys, interval=3600, first=60)

    # Автоматичний бекап раз на тиждень (604800 секунд = 7 днів)
    job_queue.run_repeating(auto_backup, interval=604800, first=3600)
    logger.info("Background job for checking deadlines scheduled (every 1 hour)")
    
    # Обработчик процесса оценки с persistence
    rating_conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(handle_rating_button, pattern='^(rate|absent)_')],
        states={
            WAITING_FOR_INTEREST: [CallbackQueryHandler(handle_interest_rating, pattern='^interest_')],
            WAITING_FOR_RELEVANCE: [CallbackQueryHandler(handle_relevance_rating, pattern='^relevance_')],
            WAITING_FOR_SPIRITUAL: [CallbackQueryHandler(handle_spiritual_rating, pattern='^spiritual_')],
            WAITING_FOR_FEEDBACK: [
                CallbackQueryHandler(handle_feedback_choice, pattern='^feedback_(yes|no)$'),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_feedback_text)
            ],
        },
        fallbacks=[
            CommandHandler('start', start),
            CallbackQueryHandler(handle_rating_button, pattern='^(rate|absent)_'),
        ],
        per_message=False,
        name="rating_conversation",
        persistent=True,
    )
    
    # Регистрируем обработчики
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", admin_help))
    application.add_handler(CommandHandler("pending", admin_pending))
    application.add_handler(CommandHandler("remove", admin_remove))
    application.add_handler(CommandHandler("start_survey", admin_start_survey))
    application.add_handler(CommandHandler("close_survey", admin_close_survey))
    application.add_handler(CommandHandler("stats", admin_stats))
    application.add_handler(CommandHandler("ratings", admin_ratings))
    application.add_handler(CommandHandler("graph", admin_graph))
    application.add_handler(CommandHandler("export_db", admin_export_db))
    application.add_handler(CommandHandler("export_excel", admin_export_excel))
    application.add_handler(CallbackQueryHandler(handle_approval, pattern='^(approve|reject|remove)_'))
    application.add_handler(rating_conv_handler)
    
    # Запускаем бота
    logger.info("Bot started!")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == '__main__':
    main()
